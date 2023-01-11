import colorama

def progress_bar(progress, total, color=colorama.Fore.MAGENTA):
    percent = 100 * ((progress) / float(total-1))
    bar = '█' * (int(percent/2)) + ' ' * (50 - (int(percent/2)))
    if progress == 1:
        print('\n')
    elif progress+1 != total:
        print(color + f"\r >>> |{bar}| {percent:.2f}%", end='\r')
    else:
        print(colorama.Fore.LIGHTGREEN_EX + f"\r >>> |{bar}| {percent:.2f}%", end='\r')
        print(colorama.Fore.RESET, '\n\n')

def applyColor(text, text_format = 1, text_color = 0, background_color = 0):
    """
    text format code:
    - 1 - None
    - 2 - Bold
    - 3 - Italic
    - 4 - Underline
    - 7 - Negative

    color code:
    - 0    - white
    - 1    - red
    - 2    - green
    - 3    - yellow
    - 4    - blue
    - 5    - purple
    - 6    - light blue
    - 7    - gray
    """
    return f"\033[{str(text_format)};{'3'+str(text_color)};{'4'+str(background_color)}m{text}\033[m"

def loading_dots():
    dot = "-"
    idx2 = 1
    side = '>'
    while True:
        yield dot
        if side == '>':
            idx2 += 1
            if idx2==5:    
                dot += "-"
                idx2 = 1
            if dot == "----":
                dot = 'ˍ'
                side = '<'
        if side == '<':
            idx2 += 1
            if idx2==5:    
                dot += "ˍ"
                idx2 = 1
            if dot == "ˍˍˍˍ":
                dot = '-'
                side = '>'

def loading_circle():
    simbol = {1:'◜', 2:'◝', 3:'◞', 4:'◟'}
    idx = 1
    while True:
        yield simbol[idx]
        idx += 1
        if idx == 5: idx = 1

def check_packages():
    import importlib.util
    from time import sleep
    import os

    package_names = ['pip', 'pandas', 'selenium', 'webdriver_manager', 'openpyxl', 'win32com', 'jinja2', 'pyautogui', 'PySimpleGUI']

    print('\nChecando dependências...\n')

    print("=======================")
    for package in package_names.copy():
        spec = importlib.util.find_spec(package)
        if spec is not None:
            package_names.remove(package)
            print(f'{package:<20} OK')
        else:
            print(f'{package:<20} X')
    print("=======================")

    if 'pip' in package_names:
        os.system('python -m ensurepip')

    if package_names:
        print('\nInstalando dependências...\n')
        for i in range(1,4):
            print(abs(4-i))
            sleep(1)
        os.system(f"pip install {' '.join(package_names)}")

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# ao inserir, tratar as colunas como números iniciando em 1
def modify(ws, index, on_horizontal_direction=False, hOrientation='left', vOrientation='top', wrap_text=False, row_space=None, col_space=None, currency=False, border=True, 
    font_name='Calibri', font_size='10', font_color='000000', font_italic=False, font_bold=False, cell_color=None):
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    if type(index)==int:
        index = [index]
    row_count = ws.max_row
    column_count = ws.max_column
    for vector in index:
        if on_horizontal_direction:
            for col in range(1, column_count+1):
                currentCell = ws[f'{get_column_letter(col)}{vector}']
                currentCell.alignment = Alignment(horizontal = hOrientation, vertical = vOrientation, wrapText=wrap_text)
                if border==True:
                    currentCell.border = thin_border
                ft = Font(name=font_name, size=font_size, color=font_color, italic=font_italic, bold=font_bold)
                currentCell.font = ft
                if cell_color is not None:
                    currentCell.fill = PatternFill("solid", start_color=cell_color)
                if col_space is not None:
                    letter = get_column_letter(col)
                    ws.column_dimensions[letter].width = col_space
            if row_space is not None:
                ws.row_dimensions[vector].height = row_space
        else:
            for row in range(2, row_count+1):
                currentCell = ws[f'{get_column_letter(vector)}{row}']
                currentCell.alignment = Alignment(horizontal = hOrientation, vertical = vOrientation, wrapText=wrap_text)
                if currency==True:
                    currentCell.number_format = 'R$ #,##0.00'
                if border==True:
                    currentCell.border = thin_border
                ft = Font(name=font_name, size=font_size, color=font_color, italic=font_italic, bold=font_bold)
                currentCell.font = ft
                if cell_color is not None:
                    currentCell.fill = PatternFill("solid", start_color=cell_color)
                if row_space is not None:
                    ws.row_dimensions[row].height = row_space
            if col_space is not None:
                letter = get_column_letter(vector)
                ws.column_dimensions[letter].width = col_space
    return ws